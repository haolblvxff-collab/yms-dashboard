"""
YMS Process Flow Import — Streamlit Page
上传 Excel Runcard 模板，解析并导入到 Process_Flow 表
"""
import streamlit as st
import pandas as pd
import sqlite3
import openpyxl
import os
import io
import re
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(__file__), "yms_data.db")


def parse_runcard_excel(file_bytes):
    """解析上传的 Excel 文件，返回 product info + process steps"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    
    result = {
        "product_id": "",
        "flow_id": "",
        "mask_family": "",
        "steps": [],
        "errors": [],
        "sheet_names": wb.sheetnames,
    }
    
    # Parse Check list for product info
    if "Check list" in wb.sheetnames:
        ws = wb["Check list"]
        for row in ws.iter_rows(min_row=5, max_row=20, values_only=True):
            for j, val in enumerate(row):
                if val is None:
                    continue
                s = str(val).strip()
                if s.startswith("FQ") and len(s) > 8:
                    result["product_id"] = s
                if s.startswith("FR") and len(s) > 8:
                    result["flow_id"] = s
                if "EF" in s and len(s) >= 5 and s[2:].replace("0", "").isalnum():
                    if result["mask_family"] == "":
                        result["mask_family"] = s
        if not result["product_id"]:
            # Try different positions
            for row in ws.iter_rows(min_row=5, max_row=15, values_only=True):
                for j, val in enumerate(row):
                    if val and 'FQ' in str(val) and len(str(val).strip()) >= 10:
                        result["product_id"] = str(val).strip()
                        break
    
    # Parse Runcard-template
    template_sheet = "Runcard-template"
    if template_sheet not in wb.sheetnames:
        result["errors"].append(f"未找到 '{template_sheet}' 工作表")
        return result
    
    ws = wb[template_sheet]
    
    # Find header row (look for "工艺流程号" in first 10 rows)
    header_row = None
    for r in range(1, min(ws.max_row + 1, 12)):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and "工艺流程号" in str(val):
                header_row = r
                break
        if header_row:
            break
    
    if header_row is None:
        result["errors"].append("未找到表头行（含'工艺流程号'列）")
        return result
    
    # Read headers
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        headers.append(str(val).strip() if val else f"Col{c}")
    
    # Parse data rows
    step_count = 0
    for r in range(header_row + 1, ws.max_row + 1):
        row_vals = []
        for c in range(1, min(len(headers) + 1, ws.max_column + 1)):
            val = ws.cell(row=r, column=c).value
            row_vals.append(str(val).strip() if val is not None else "")
        
        # Check if this is a valid step row
        route = row_vals[1] if len(row_vals) > 1 else ""
        step_id = row_vals[2] if len(row_vals) > 2 else ""
        desc = row_vals[3] if len(row_vals) > 3 else ""
        
        # Skip empty rows, header-like rows, rework-only rows
        if not route or route == "路径号(*)":
            continue
        if route.startswith("ETD_") and len(route) > 10:
            continue
        if not step_id or not desc:
            continue
        
        # Skip non-process rows (pure descriptions like "在此下方继续添加")
        if not re.match(r'^\d{6}$', str(step_id)):
            continue
        
        step_count += 1
        step = {
            "flow_id": row_vals[0] if len(row_vals) > 0 else "",
            "route": route,
            "step_id": step_id,
            "description": desc,
            "target": row_vals[4] if len(row_vals) > 4 else "",
            "spc_job_id": row_vals[5] if len(row_vals) > 5 else "",
            "recipe_id": row_vals[6] if len(row_vals) > 6 else "",
            "equip_group": row_vals[7] if len(row_vals) > 7 else "",
            "equip_id": row_vals[8] if len(row_vals) > 8 else "",
            "mask_family": row_vals[9] if len(row_vals) > 9 else "",
            "mask_id": row_vals[10] if len(row_vals) > 10 else "",
            "edc_param_set": row_vals[11] if len(row_vals) > 11 else "",
            "qtime": row_vals[12] if len(row_vals) > 12 else "",
            "qtime_id": row_vals[13] if len(row_vals) > 13 else "",
            "rework_flow": row_vals[14] if len(row_vals) > 14 else "",
            "rework_step": row_vals[15] if len(row_vals) > 15 else "",
            "route_branch": row_vals[16] if len(row_vals) > 16 else "",
            "step_branch": row_vals[17] if len(row_vals) > 17 else "",
            "operation_hint": row_vals[18] if len(row_vals) > 18 else "",
            "operation_type": row_vals[19] if len(row_vals) > 19 else "",
        }
        result["steps"].append(step)
    
    # Get flow_id from data if not found in Check list
    if not result["flow_id"] and result["steps"]:
        for s in result["steps"]:
            if s["flow_id"] and s["flow_id"].startswith("FR"):
                result["flow_id"] = s["flow_id"]
                break
    
    if step_count == 0:
        result["errors"].append("未解析到有效工步数据")
    
    return result


def import_to_db(steps, flow_id, product_id):
    """将解析的工步写入 Process_Flow 表，返回导入数量"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Ensure table exists
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Process_Flow (
            ID INTEGER PRIMARY KEY AUTOINCREMENT,
            Flow_ID TEXT,
            Route_ID TEXT,
            Step_ID TEXT,
            Step_Description TEXT,
            Target TEXT,
            Recipe_ID TEXT,
            Equipment_Group TEXT,
            Equipment_ID TEXT,
            Mask_Family TEXT,
            Mask_ID TEXT,
            EDC_Param_Set TEXT,
            QTime TEXT,
            QTime_ID TEXT,
            Rework_Flow TEXT,
            Rework_Step TEXT,
            Operation_Hint TEXT,
            Seq_No INTEGER
        )
    ''')
    
    # If flow_id specified, delete old data for this flow
    if flow_id:
        cursor.execute("DELETE FROM Process_Flow WHERE Flow_ID = ?", (flow_id,))
    
    inserted = 0
    for i, s in enumerate(steps):
        fid = s.get("flow_id", "") or flow_id
        recipe = s.get("recipe_id", "")
        if recipe == "NA":
            recipe = ""
        equip = s.get("equip_group", "")
        if equip == "NA":
            equip = ""
        
        cursor.execute('''
            INSERT INTO Process_Flow 
            (Flow_ID, Route_ID, Step_ID, Step_Description, Target, Recipe_ID, 
             Equipment_Group, Equipment_ID, Mask_Family, Mask_ID,
             EDC_Param_Set, QTime, QTime_ID, Rework_Flow, Rework_Step, Operation_Hint, Seq_No)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            fid, s["route"], s["step_id"], s["description"], s["target"],
            recipe, equip, s.get("equip_id", ""),
            s.get("mask_family", ""), s.get("mask_id", ""),
            s.get("edc_param_set", ""), s.get("qtime", ""), s.get("qtime_id", ""),
            s.get("rework_flow", ""), s.get("rework_step", ""),
            s.get("operation_hint", ""),
            i + 1
        ))
        inserted += 1
    
    conn.commit()
    conn.close()
    return inserted


def render_process_flow_page(db_run_query):
    """Streamlit 页面入口"""
    st.title("📋 Process Flow Import")
    st.markdown("上传 Excel Runcard 模板，导入工艺流程数据")
    
    # ===== TAB 1: Import =====
    tab1, tab2 = st.tabs(["📥 Import Runcard", "📊 Current Process Flow"])
    
    with tab1:
        st.markdown("### 1. 下载模板")
        
        template_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), 
            "templates", "runcard_template.xlsx"
        )
        if os.path.exists(template_path):
            with open(template_path, "rb") as f:
                st.download_button(
                    "📥 下载 Runcard 模板 (.xlsx)",
                    f,
                    "runcard_template.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        else:
            st.warning("模板文件未找到，请参考下方说明自行创建。")
        
        with st.expander("📖 模板格式说明"):
            st.markdown("""
            **必填列**（`Runcard-template` sheet）：
            - `工艺流程号(*)` — 工艺版本标识，如 `FR003WM2A001E001`
            - `路径号(*)` — 工艺路径简称，如 `S1`, `U1`, `V1`, `M1`
            - `步骤号(*)` — 6位步骤编号，如 `020040`
            - `步骤描述(*)` — 工步描述文字
            - `RECIPE ID(*)` — Recipe 名称，无则填 `NA`
            - `设备组ID(*)` — 设备组代号，无则填 `NA`
            
            **选填列**：`TARGET`, `SPC JOB ID`, `EDC参数集ID`, `Q-TIME` 等
            
            **Check list sheet** 填写产品号和工艺流程号（可选，用于自动识别）。
            """)
        
        st.markdown("### 2. 上传文件")
        uploaded = st.file_uploader(
            "选择 .xlsx 文件",
            type=["xlsx"],
            help="支持标准 Runcard 格式（与 B3RxRuncard.xlsx 兼容）"
        )
        
        if uploaded:
            file_bytes = uploaded.read()
            
            with st.spinner("解析中..."):
                parsed = parse_runcard_excel(file_bytes)
            
            if parsed["errors"]:
                for e in parsed["errors"]:
                    st.error(e)
            
            if parsed["steps"]:
                st.success(f"✅ 解析成功：{len(parsed['steps'])} 个工步")
                
                # Product info
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("产品号", parsed["product_id"] or "未识别")
                with col2:
                    st.metric("工艺流程号", parsed["flow_id"] or "从数据中识别")
                with col3:
                    st.metric("光罩族", parsed["mask_family"] or "未识别")
                
                # Route summary
                routes = {}
                for s in parsed["steps"]:
                    routes[s["route"]] = routes.get(s["route"], 0) + 1
                
                st.markdown(f"**路径分布** ({len(routes)} 条)：")
                route_summary = " | ".join(f"{r}({c})" for r, c in sorted(routes.items()))
                st.text(route_summary)
                
                # Preview
                with st.expander("🔍 预览前20行"):
                    preview_data = []
                    for s in parsed["steps"][:20]:
                        preview_data.append({
                            "路径": s["route"],
                            "步骤号": s["step_id"],
                            "步骤描述": s["description"][:40],
                            "Recipe": s["recipe_id"][:25],
                            "设备组": s["equip_group"][:8],
                            "EDC": s["edc_param_set"][:20],
                        })
                    st.dataframe(pd.DataFrame(preview_data), use_container_width=True)
                
                # Import button
                st.markdown("### 3. 导入数据库")
                
                flow_id_input = st.text_input(
                    "工艺流程号（确认）",
                    value=parsed["flow_id"] or "",
                    help="将覆盖此工艺流程号下的旧数据"
                )
                
                if st.button("🚀 导入到 Process_Flow 表", type="primary", use_container_width=True):
                    with st.spinner("导入中..."):
                        count = import_to_db(
                            parsed["steps"],
                            flow_id_input or parsed["flow_id"],
                            parsed["product_id"]
                        )
                    st.success(f"✅ 成功导入 {count} 个工步到 Process_Flow 表")
                    st.info("刷新页面或切换到 'Current Process Flow' 标签查看结果")
                    st.balloons()
    
    # ===== TAB 2: Current Data =====
    with tab2:
        try:
            df = db_run_query("""
                SELECT Flow_ID, Route_ID, Step_ID, Step_Description, Recipe_ID, 
                       Equipment_Group, EDC_Param_Set
                FROM Process_Flow 
                ORDER BY Seq_No
            """)
            
            if not df.empty:
                # Flow selector
                flows = df['Flow_ID'].unique().tolist()
                selected_flow = st.selectbox("工艺流程号", flows)
                
                flow_df = df[df['Flow_ID'] == selected_flow]
                
                # Stats
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("总工步", len(flow_df))
                with col2:
                    routes_n = flow_df['Route_ID'].nunique()
                    st.metric("路径数", routes_n)
                with col3:
                    recipes_n = flow_df[flow_df['Recipe_ID'] != '']['Recipe_ID'].nunique()
                    st.metric("Recipes", recipes_n)
                with col4:
                    equips_n = flow_df[flow_df['Equipment_Group'] != '']['Equipment_Group'].nunique()
                    st.metric("设备组", equips_n)
                
                st.dataframe(
                    flow_df[['Route_ID', 'Step_ID', 'Step_Description', 'Recipe_ID', 'Equipment_Group']],
                    use_container_width=True,
                    height=400
                )
                
                # Export
                csv = flow_df.to_csv(index=False).encode('utf-8')
                st.download_button("📥 导出 CSV", csv, f"process_flow_{selected_flow}.csv", "text/csv")
            else:
                st.info("暂无工艺流程数据。请先导入。")
        except Exception as e:
            st.error(f"数据库查询失败：{e}")
